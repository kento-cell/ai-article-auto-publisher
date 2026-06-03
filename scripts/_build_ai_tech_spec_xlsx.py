"""Generate a detailed Excel workbook documenting the AI internals of the
article auto-publisher system.

One-shot doc generator (2026-06-03). Values are transcribed from the
codebase as explored; code line numbers may drift over time. Runtime
values (from .env) are flagged separately from code defaults.

Run: py scripts/_build_ai_tech_spec_xlsx.py
Out: docs/ai_technical_spec_20260603.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if sys.platform == "win32" and hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _ai_spec_diagrams as dg  # noqa: E402

_REPO = Path(__file__).resolve().parent.parent
OUT = _REPO / "docs" / "ai_technical_spec_20260603.xlsx"
_ASSETS = _REPO / "docs" / "_spec_assets"

# ---- palette / styles ------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5395"
LIGHT = "D9E1F2"
ZEBRA = "F2F5FB"
ACCENT = "C55A11"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Meiryo", size=15, bold=True, color=WHITE)
DESC_FONT = Font(name="Meiryo", size=9, italic=True, color="404040")
HEAD_FONT = Font(name="Meiryo", size=10, bold=True, color=WHITE)
CELL_FONT = Font(name="Meiryo", size=10, color="222222")
KEY_FONT = Font(name="Meiryo", size=10, bold=True, color="1F3864")

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FILL = PatternFill("solid", fgColor=BLUE)
ZEBRA_FILL = PatternFill("solid", fgColor=ZEBRA)
KEY_FILL = PatternFill("solid", fgColor=LIGHT)

TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

_thin = Side(style="thin", color="B4C6E7")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def add_sheet(wb, name, title, description, columns, rows, widths):
    """Add one formatted sheet. columns=header list, rows=list of lists."""
    ws = wb.create_sheet(title=name)
    ncol = len(columns)
    last_col = get_column_letter(ncol)

    # Title banner (row 1)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Description (row 2)
    ws.merge_cells(f"A2:{last_col}2")
    d = ws["A2"]
    d.value = description
    d.font = DESC_FONT
    d.alignment = TOP
    ws.row_dimensions[2].height = 34 if len(description) > 70 else 20

    # Header (row 3)
    for j, head in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=j, value=head)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    # Body
    for i, row in enumerate(rows, start=4):
        for j in range(1, ncol + 1):
            val = row[j - 1] if j - 1 < len(row) else ""
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = KEY_FONT if (ncol == 2 and j == 1) else CELL_FONT
            cell.alignment = TOP
            cell.border = BORDER
            if ncol == 2 and j == 1:
                cell.fill = KEY_FILL
            elif (i % 2) == 0:
                cell.fill = ZEBRA_FILL

    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    return ws


def add_image_sheet(wb, name, banner, png_path, max_w=1180):
    ws = wb.create_sheet(title=name)
    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value = banner
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    img = XLImage(str(png_path))
    if img.width and img.width > max_w:
        ratio = max_w / img.width
        img.width = max_w
        img.height = int(img.height * ratio)
    ws.add_image(img, "A3")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)

    # ---- 00 表紙 / 概要 ---------------------------------------------
    add_sheet(
        wb, "00_概要",
        "AI記事自動生成システム — 技術仕様書 (AI内部詳細)",
        "対象リポジトリ: ai-article-auto-publisher / 作成: 2026-06-03 / "
        "値はコード探索時点。行番号は将来ドリフトし得る。ランタイム値は .env 由来、"
        "別途コード default と併記。",
        ["項目", "内容"],
        [
            ["目的", "Zenn(技術プレミアム) と note(一般バズ) に高品質記事を自動生成・"
             "投稿。5専門エージェントの議論で品質担保。"],
            ["最上位ルール", "タイトルは攻めてよい(グレーギリギリOK)。ただし大々的に謳った以上、"
             "中身は必ず濃く。タイトル負け=読者への裏切り=絶対禁止。"],
            ["AIスタック", "ローカルLLM (Ollama / Gemma) + RAG (chromadb + multilingual-e5-base) "
             "+ 画像生成 (ChatGPT image tool via Brave CDP) + 規則ベース スコアリング。"],
            ["パイプライン", "収集 → トレンドスコア → 構成パターン選択 → LLM生成 → 画像生成 → "
             "客観スコア(足切り) → 主観スコア(LLM) → 集約判定(A/B/C) → Sheets/store → publish。"],
            ["品質判定", "総合 = min(客観最低, 主観平均)。客観Cが1つでもあれば総合Cで自動却下。"],
            ["人間承認", "Sheets で ⏳承認待ち → ✅承認 を人が付けて初めて publish 対象。"],
            ["主要モジュール", "collectors/ generators/ publishers/ utils/ — 詳細は別シート。"],
            ["シート一覧", "01パイプライン 02エージェント 03LLM 04プロンプト 05構成パターン "
             "06客観スコア 07主観スコア 08集約 09RAG 10ハルシ対策 11画像生成 12収集トレンド "
             "13価格ハッシュタグ 14主要パラメータ 15ファイル対応表。"],
        ],
        [22, 95],
    )

    # ---- 図解シート (PIL生成図を埋め込み) ----------------------------
    for name, title, png in dg.generate_all(_ASSETS):
        banner = f"{name.split('_')[0]}  {title}"
        add_image_sheet(wb, name, banner, png)

    # ---- 01 パイプライン --------------------------------------------
    add_sheet(
        wb, "01_パイプライン",
        "処理パイプライン (--generate / --publish)",
        "main.py が全体を駆動。--generate=生成系 Phase1-3、--publish=Phase4。",
        ["#", "フェーズ", "処理内容", "主担当 / ファイル"],
        [
            ["1", "収集 (Phase1)", "arXiv + RSS(日/韓) + Reddit + Bluesky + Google Trends "
             "+ knowledge_topics を取得", "collectors/*"],
            ["2", "トレンドスコア", "recency0.40 + social0.35 + authority0.25 で 0-100 ランク付け",
             "collectors/trend_detector.py"],
            ["3", "構成パターン選択", "title+source からキーワード/カテゴリで listicle/trend_report/"
             "howto/tutorial/deep_dive 等を自動選択", "config/prompts.yaml structure_selection"],
            ["4", "LLM記事生成", "構成パターンをプロンプトに注入し Gemma で本文生成 "
             "(Zenn 5000-7000字 / note 2800-3500字)", "generators/local_llm.py + prompts.yaml"],
            ["5", "サニタイズ", "プレースホルダ/AI開示行をスコアリング前に除去",
             "generators/content_sanitizer.py"],
            ["6", "画像生成", "ChatGPT image tool で cover+inline、vision-eval で自己採点、"
             "失敗時 Pollinations→Unsplash", "generators/chatgpt_image_generator.py ほか"],
            ["7", "客観スコア(足切り)", "Tier比率/引用/視覚/禁止句/チェーン店等を A/B/C 判定。"
             "C が1つでもあれば不合格", "generators/objective_scorer.py"],
            ["8", "主観スコア(LLM)", "独自性/正確性/可読性/引き込み/タイトル回収 を根拠付き A/B/C",
             "generators/subjective_evaluator.py"],
            ["9", "集約判定", "総合 = min(客観最低, 主観平均)。C は自動却下、それ以外 Sheets 登録",
             "generators/score_aggregator.py"],
            ["10", "登録/通知", "ArticleStore(JSON) + Sheets(⏳承認待ち) + Gmail 通知",
             "utils/article_store.py, sheets_manager.py"],
            ["11", "publish (Phase4)", "Sheets ✅承認 を取得 → cadence cap/dup/deny → "
             "Zenn(git push/scrap) / note(価格決定+Selenium投稿) → Slack/Gmail",
             "main.py::publish_approved"],
        ],
        [4, 18, 62, 34],
    )

    # ---- 02 エージェント --------------------------------------------
    add_sheet(
        wb, "02_AIエージェント",
        "5専門エージェント (議論ベース・会議型)",
        "Researcher の調査結果が土台。全ての信頼性はそこから派生。Critic は常に否定から入る。"
        "出典: AGENTS.md, .claude/skills/<agent>/SKILL.md。",
        ["エージェント", "役割", "主な責務 / 出力", "重要ルール"],
        [
            ["Researcher\n(調査)", "深い調査・事実確認・ソース信頼性評価",
             "verified_facts / data_points / counterarguments / source_list / unresolved。"
             "ソース4 Tier 評価(1学術/官公→4匿名)", "主要主張は Tier1-2 を最低3独立ソース。"
             "統計は一次データまで遡及。未検証を事実扱いしない。"],
            ["Strategist\n(戦略)", "トピック選定・差別化角度・ターゲット設定",
             "angle / persona / structure(Hook→Value→CTA) / differentiation / "
             "platform_specific / reject_reason", "全記事に『なぜ今これか』。差別化できなければ"
             "REJECT。2週間以内の重複トピック禁止。"],
            ["Writer\n(執筆)", "リッチ整形・視覚要素・著作権安全画像で本文ドラフト",
             "構成ブリーフを実行。視覚要素 最低3 (図/表/画像)、tech は Mermaid 最低1、比較表最低1、"
             "コールアウト最低2", "画像は CC0/Unsplash/Pexels のみ + 帰属必須。H1は本文に置かない。"],
            ["Critic\n(批評)", "否定起点でドラフトを Researcher 根拠と照合",
             "[Critic Round N] 形式で 否定/指摘→修正案、未解消数・新規数・収束判断を毎回提示",
             "肯定しない(『指摘なし=収束可』のみ)。Writer↔Critic 最大2(例外3)round。"
             "薄い記事は他指標が良くても却下。"],
            ["Coordinator\n(調整)", "議論進行・収束判断・スコア集約・ユーザー提示",
             "ラウンド進行管理、客観(機械測定)+主観(議論由来)を集約、Sheets 行を生成",
             "収束条件: Critic未解消0 + 客観Cなし + 未検証主張なし + 証拠Lv B以上。"],
        ],
        [16, 30, 56, 40],
    )

    # ---- 03 LLM 構成 ------------------------------------------------
    add_sheet(
        wb, "03_LLM構成",
        "ローカルLLM (Ollama) 構成",
        "ランタイム(.env)とコード default を区別。2026-05-15 に Writer/Scorer を "
        "gemma3:12b → gemma4:e4b へ切替(長文+密度保持、速度向上)。",
        ["項目", "値", "出所 / 備考"],
        [
            ["実行モデル (Writer)", "gemma4:e4b", ".env LLM_MODEL_WRITER (ランタイム)"],
            ["実行モデル (Scorer)", "gemma4:e4b", ".env LLM_MODEL_SCORER (ランタイム)"],
            ["コード default モデル", "gemma3:12b", "generators/local_llm.py DEFAULT_MODEL"],
            ["タスク別モデル", "writer/scorer/summarizer/hashtag/regenerator",
             "generators/llm_config.py。LLM_MODEL_<TASK> で個別上書き可"],
            ["Ollama URL", "http://localhost:11434", ".env OLLAMA_API_URL。loopback 限定(セキュリティ)"],
            ["タイムアウト", "900 秒 (15分)", "Gemma3:12b は 5000-7000字で 6-12分。"
             "旧300秒上限がバッチ全滅を起こした"],
            ["温度 (記事生成)", "0.7", "創造性重視"],
            ["温度 (批評/評価)", "0.3", "厳密・精密"],
            ["温度 (要約)", "0.5", "バランス"],
            ["温度 (コード生成)", "0.2", "local_llm.generate_code"],
            ["num_ctx (文脈長)", "16384", "2026-05-15 に 8192→16384。Writerプロンプト~8900tok+"
             "出力で 8192 が末尾切れ→引用規則/タイトル節喪失していた"],
            ["keep_alive", "30m", "モデルをウォーム保持し呼び出し間の再ロードを回避"],
            ["プロンプト最大長", "16000 字", "Codex フォールバック時の上限"],
            ["フォールバック", "OpenAI Codex CLI", "Ollama 失敗時。LLM_CODEX_FALLBACK=false で無効化。"
             "codex exec --sandbox read-only"],
            ["週次トークン上限", "2,000,000 tokens", "config/settings.yaml。80%(160万)でアラート、50%バッファ"],
        ],
        [22, 40, 58],
    )

    # ---- 04 プロンプト設計 ------------------------------------------
    add_sheet(
        wb, "04_プロンプト設計",
        "プロンプト設計 (config/prompts.yaml, 約991行)",
        "トップキー: zenn_article_prompt / note_article_prompt / quality_evaluation_prompt "
        "/ article_structures / structure_selection。AI臭を消す・タイトル回収を最重視。",
        ["プロンプト / 規則", "要点"],
        [
            ["zenn_article_prompt", "技術記事。引用ブロックは > 内に URL+取得日 必須。数値捏造・"
             "プレースホルダ(〇〇mg等)・伏字・仮名・AI開示は1件で即不合格。見出しは ## (太字見出し禁止)。"],
            ["note_article_prompt", "一般/トレンド。バズ最適化。薄い汎論禁止、記事ごと初見の具体事実5件以上。"
             "密度>文字数。店舗系は STORE_BLOCK + Google Places 検証、チェーン店禁止、住所/時間/価格は書かない。"],
            ["引用ブロック形式", "> \"引用\" / > 出典: 著者. \"タイトル\" / > {url} / > (取得日: YYYY年MM月DD日)"],
            ["整合性ルール(即不合格)", "ソースに無い数値の捏造禁止 / プレースホルダ禁止(2026-05-26) / "
             "実在名称必須・伏字仮名禁止(2026-05-27) / 借用画像は price=¥0 強制"],
            ["AI臭を消す", "絵文字(😀🎉)禁止=AI記事の指紋。顔文字 (^^)(>_<) は2-4個推奨。"
             "ですます+砕けた崩し(〜ですよね/ぶっちゃけ)。"],
            ["書き出し", "一人称/読者巻き込み/告白型が人気。『本記事では〜解説します』等の教科書調は禁止。"],
            ["タイトル規則", "【】開始(TOP30の80%)。35-45字(最大70)。数字必須。絵文字0%。"
             "煽り+具体+希少。タイトルで匂わせたら本文で必ず回収。"],
            ["必殺テンプレート5選", "例:【殿堂入り記事】【警告】〇〇は嘘だった。データが証明する「本当の〇〇」"],
            ["quality_evaluation_prompt", "LLM品質評価。originality/accuracy/readability/citation/"
             "practicality を各0-10、total0-50、pass/should_regenerate を JSON で。"],
            ["末尾 免責表記(note)", "店舗系は『Google Maps公開データ等をもとに構成。来店前に公式確認を』"
             "の免責を必ず付与。"],
        ],
        [26, 90],
    )

    # ---- 05 構成パターン --------------------------------------------
    add_sheet(
        wb, "05_構成パターン",
        "構成パターン (article_structures + structure_selection)",
        "8テンプレートを回転させ過学習を防止。title+source の部分一致でルーティング。",
        ["テンプレート", "用途", "アウトライン / 選択トリガー"],
        [
            ["standard", "汎用・安全", "導入→概要→詳細→まとめ。default フォールバック"],
            ["tutorial", "入門/始め方", "目標→前提→Step1/2/3→トラブルシュート。kw: 入門/始め方/tutorial(zenn)"],
            ["comparison", "ツール/製品比較", "背景→候補→評価基準→比較表→ユースケース→結論。kw: 比較/vs/選び方"],
            ["deep_dive", "研究/分析", "なぜ今重要→基礎→深掘り→事例→落とし穴→筆者見解→参考。kw: 仕組み/arxiv/paper"],
            ["qa", "FAQ", "導入質問→Q1/2/3→統合→さらに読む"],
            ["listicle", "『5選！』", "なぜこのリスト→選定基準→No.1/2/3→次点→CTA。note gourmet/cafe"],
            ["trend_report", "速報/最新性", "何が起きた→なぜ今→現地反応→日本影響→アクセス→予測。"
             "note 韓国/kpop/kbeauty/速報"],
            ["howto", "プロの技 (コーヒー/美容)", "学べること→必要道具→Step→失敗対処→プロのコツ。"
             "note beauty/selfcare/バリスタ"],
        ],
        [16, 24, 76],
    )

    # ---- 06 客観スコアリング ----------------------------------------
    add_sheet(
        wb, "06_客観スコアリング",
        "客観スコアリング (足切り) — generators/objective_scorer.py",
        "規則ベース(LLM不使用)。C が1つでもあれば不合格。note/arXiv には一部緩和あり。",
        ["指標", "A", "B", "C", "備考"],
        [
            ["evidence_level\n(Tier1-2率)", "≥80%", "60-79%", "<60%", "Tier1-2ソース比率"],
            ["citation_count\n(引用数)", "≥5", "2-4", "≤1", "引用ブロック+参考リンク。note/arXiv は非ブロッキング"],
            ["citation_format\n(引用形式)", "URL+日付100%", "≥50% or 1件grounded", "引用なし",
             "arXiv は URL自動クレジット"],
            ["visual_count\n(視覚要素)", "≥5", "1-4", "0", "画像+Mermaid+表+コード。2026-05-14 緩和(Bは1から)"],
            ["word_count\n(文字数)", "2,200-3,500", "1,700-8,000", "範囲外",
             "note最盛帯。accept_max は gemma4 長文対応で 5,500→8,000 に緩和"],
            ["title_fulfillment\n(タイトル回収)", "回収十分", "中間", "タイトル負け",
             "title_fulfillment_scorer。『5選』で3件等の機械的看板倒れを検出"],
            ["forbidden_phrases\n(禁止句)", "0違反", "—", "違反あり",
             "settings.yaml の40+正規表現。note は構造テンプレ系のみ非ブロッキング"],
            ["heading_structure\n(見出し)", "Pass", "—", "Fail", "H2 最低2、本文 H1 は0"],
            ["chain_stores\n(チェーン店)", "Pass", "—", "Fail", "25件ブラックリスト(グルメ記事)"],
            ["[加点] trend_alignment", "トレンド一致", "(床)", "—", "非ブロッキング加点"],
            ["[加点] first_hand_experience", "E-E-A-T 3マーカー以上", "(床)", "—", "非ブロッキング加点"],
        ],
        [22, 16, 18, 14, 44],
    )

    # ---- 07 主観スコアリング ----------------------------------------
    add_sheet(
        wb, "07_主観スコアリング",
        "主観スコアリング (LLM評価) — generators/subjective_evaluator.py",
        "5次元を根拠必須で A/B/C。各次元20字以上の reason 必須。JSON 解析失敗時は全Bに graceful fallback。",
        ["次元", "A", "B(床)", "C", "備考"],
        [
            ["originality\n(独自性)", "真に新規な分析/統合", "調査/ニュース要約+筆者見解",
             "純粋な機械翻訳・無相互参照・無意見", "筆者見解があれば B 床"],
            ["accuracy\n(正確性)", "全主張が正確・出典付き", "概ね正確(軽微な隙はOK)",
             "未検証断定・ブリーフ矛盾", "RAG hallucination_warnings で C 降格(本文に実在確認時)"],
            ["readability\n(可読性)", "明快な構造・スキャン可", "良い流れ(多少の難はOK)",
             "平坦・無構造", ""],
            ["engagement\n(引き込み)", "強いフック+物語+示唆", "フック+構造+示唆あり",
             "物語性なし・フックなし", "tech は床B"],
            ["title_fulfillment\n(タイトル回収)", "具体/実数/固有名/引用で回収",
             "話題には触れるが『落ち』が弱い", "看板倒れ(タイトル負け)=ハード却下",
             "最重要。C は総合即却下"],
        ],
        [20, 30, 30, 30, 30],
    )

    # ---- 08 スコア集約 ----------------------------------------------
    add_sheet(
        wb, "08_スコア集約",
        "スコア集約 — generators/score_aggregator.py",
        "総合 = min(客観最低, 主観平均) の思想。C は自動却下。GRADE値 A=3/B=2/C=1。",
        ["ルール / 値", "内容"],
        [
            ["VETO(客観C)", "客観指標が1つでもC → 総合C 却下"],
            ["VETO(タイトル)", "title_fulfillment==C → 総合C 却下(『タイトル負けで却下』)"],
            ["主観C + 客観ブロッキング", "→ 総合C"],
            ["主観C(客観ブロッキングなし)", "→ 総合B で承認"],
            ["全A", "客観全A + 主観全A → 総合A(承認推奨)"],
            ["B+閾値", "主観平均 ≥ 2.0 (2026-05-14 に 2.5 から緩和) → 総合B"],
            ["GRADE値", "A=3 / B=2 / C=1"],
            ["numeric_score", "0.5×客観平均 + 0.5×主観平均。GRADE_NUMERIC A=100/B=75/C=50"],
            ["Zenn 記事閾値", "numeric_score ≥ 77.5 で記事投稿、未満は Zenn Scrap (main.py publish側)"],
            ["decision", "C=reject / B,A=approve。blocking_issues に objective:/subjective: 接頭辞付きで列挙"],
            ["Sheets ステータス", "却下=❌自動却下 / 承認候補=⏳承認待ち"],
        ],
        [26, 90],
    )

    # ---- 09 RAG -----------------------------------------------------
    add_sheet(
        wb, "09_RAG構成",
        "RAG構成 (chromadb + multilingual-e5-base)",
        "RAG_ENABLED=true (本番ON)。e5 接頭辞 passage:(索引)/query:(検索)。"
        "索引: scripts/build_rag_index.py、検索: generators/rag_retriever.py。",
        ["項目", "値 / 内容"],
        [
            ["埋め込みモデル", "intfloat/multilingual-e5-base (768次元, 約280MB)"],
            ["ベクトルDB", "chromadb (永続化: data/rag_index/)"],
            ["コレクション", "anti_patterns / successes / hallucinations / ops_incidents / "
             "generation_guides / past_articles / thumbnail_styles"],
            ["生成時RAG", "RAG_ENABLED で anti_patterns+successes を top_k3, 閾値0.55 で取得し"
             "学習ブロックを意味検索化 (main.py)"],
            ["ハルシ ガード", "RAG_HALLUCINATION_CHECK(既定ON)。hallucinations を閾値0.85で照合、"
             "multi-query 展開、reranker OFF(ハルシ精度を下げるため)"],
            ["重複検出", "past_articles コレクション(タイトル+要約300字)で類似記事を検出"],
            ["ops バナー", "ops_incidents/generation_guides を query し [ops-banner:*] を生成時/publish時に表示"],
            ["再ランカ", "RAG_RERANKER(既定ON, BAAI/bge-reranker-base)。ハルシガードでは無効"],
            ["オフライン", "HF_HUB_OFFLINE=1 / TRANSFORMERS_OFFLINE=1 (ダウンロードハング回避)"],
            ["規律", "新事象→ops_incidents.md 追記+再ingest。logic変更時は先に query して既存事象とコンフリ確認。"],
        ],
        [22, 92],
    )

    # ---- 10 ハルシネーション対策 ------------------------------------
    add_sheet(
        wb, "10_ハルシ対策",
        "ハルシネーション対策 (3層 deny + 3箇所同期)",
        "新事故時は hallucination_registry.md を正典に、deny を3箇所同期し test_hallucination_deny.py で回帰。",
        ["層 / 項目", "内容"],
        [
            ["層1: 禁止句正規表現", "config/settings.yaml に40+パターン。SNS捏造(氏のBluesky投稿)/"
             "架空イベント/プレースホルダ/伏字+業態(〇〇寿司)/記号命名/仮名/AI開示footer/接続詞スパム/"
             "em-dashスパム/段階解説。起動時に1回コンパイル。"],
            ["層2: サニタイザ", "generators/content_sanitizer.py。スコアリング前に問題行を除去。"
             "_LINE_KILL_PHRASES(ここに入力/URLは記載しません/架空のURL等)、AI開示行。"],
            ["層3: RAG ハルシガード", "Critic が hallucination_warnings を受領→本文に実在確認できれば "
             "accuracy を C 降格し該当箇所を引用。閾値0.85。"],
            ["publish時 deny", "main.py::publish_approved の _PUBLISH_DENY_PATTERNS(最終防衛線)。"
             "title/JP抽出/本文head2500/tail2500 を検査、hit で Sheets 行を ❌却下。"],
            ["正規表現安全化", "2026-05-15 catastrophic backtracking 修正。ネスト {2,} を atomic 単一量化に。"
             "接続詞は出現回数<3で短絡、>1秒のパターンはログ。"],
            ["同期3箇所", "(1) settings.yaml(正典) (2) test_hallucination_deny.py(40 deny+7 sanitizer+3 RAG) "
             "(3) main.py _PUBLISH_DENY_PATTERNS。"],
            ["保守フロー", "事故検知→registry.md追記→settings.yaml更新→test実行→build_rag_index.py再ingest。"],
        ],
        [22, 92],
    )

    # ---- 11 画像生成 ------------------------------------------------
    add_sheet(
        wb, "11_画像生成AI",
        "画像生成AI (ChatGPT image tool via Brave CDP)",
        "API課金なし。Brave を CDP attach し ChatGPT Web UI の画像ツールを Playwright で駆動。"
        "プロンプトは Gemma が日本語要約→ビジュアル化。",
        ["項目", "値 / 内容"],
        [
            ["生成手段", "ChatGPT 画像ツール(gpt-image-1.5系)を Brave+Playwright で駆動。Web UI 経由(無料枠)"],
            ["CDP attach", ".env CHATGPT_CDP_PORT=9222。既存 Brave に 127.0.0.1:9222 で attach。"
             "失敗時 launch_persistent_context(Brave完全停止要)へ fallback"],
            ["固定チャットURL", "data/chatgpt-image-chat-url.txt(share URL)。全生成を1会話に固定"],
            ["セッション方針", "画像1枚ごとに new chat → 抽出後即 soft-delete(PATCH is_visible:false)。"
             "sidebar leak 防止、失敗パスでも cleanup"],
            ["プロンプト builder", "generators/visual_prompt_builder.py。Gemma が記事を日本語2-3文(≤150字)の"
             "シーン記述に変換。<ARTICLE_META>でプロンプト注入防御"],
            ["サイズ", "landscape 16:9 1792×1024(既定) / portrait 9:16 1024×1792 / square 1:1 1024×1024"],
            ["既定スタイル", "宮崎駿/新海誠/細田守風の手描き水彩アニメ調(2026-04-28 に『スタジオジブリ風』直書きから変更)。"
             "テキスト/ロゴ/UIは描かない"],
            ["cover vs inline", "cover=クリックベイト インフォグラフィック(大きな日本語タイトル文字)。"
             "inline=テキスト控えめ水彩"],
            ["vision-eval", "生成画像を ChatGPT 自身が SCORE:1-10 で自己採点。cutoff=6(未満は1回だけ再生成)。"
             "CHATGPT_VISION_EVAL 既定ON、fail-closed"],
            ["style_preset", "generators/image_style_presets.py。kbeauty_poster(cover_styled=True, "
             "韓国美容雑誌エディトリアル実写調)。chatgpt_image_batch(style_preset=...) で指定"],
            ["フォールバック段", "ChatGPT → (USE_POLLINATIONS_FALLBACK時)Pollinations flux 1200×630 → Unsplash/Pexels"],
            ["重複ガード", "バッチ内で同一MD5(プレースホルダ黄ロゴ)を検出したら全重複を無効化し fallback"],
            ["最小バイト", "_MIN_VALID_IMAGE_BYTES=10,000。プレースホルダ混入を排除"],
            ["タイムアウト", "ナビ60s / 画像生成240s / テキスト応答120s"],
            ["借用画像", "公式SNS等の第三者画像を本文に置いた記事は paid 禁止(price=¥0 強制, main.py)"],
        ],
        [20, 94],
    )

    # ---- 12 収集・トレンド ------------------------------------------
    add_sheet(
        wb, "12_収集トレンド",
        "収集 & トレンド検出 — collectors/",
        "BaseCollector(rate_limit1.0s, timeout30s)。トレンドスコア=0-100。",
        ["項目", "値 / 内容"],
        [
            ["arXiv", "arxiv_collector.py。CS.AI/CL/LG/CV/MA/SE/CR、~50論文。authority 0.90"],
            ["Reddit", "reddit_collector.py。programming/MachineLearning/artificial/technology。"
             "authority 0.60-0.75 (※近時 403 多発)"],
            ["RSS", "rss_collector.py。40+フィード(はてな/Yahoo/Gigazine/美容誌/グルメ)。"
             "feed config に target=note/zenn"],
            ["Bluesky / Google Trends", "bluesky_collector.py / google_trends_collector.py。"
             "social signals、note向けトレンド語"],
            ["knowledge_topics", "knowledge_topics_collector.py。常緑シード(k_beauty/hidden_gourmet/"
             "coffee_barista/self_improvement/ai_sidejob)。trend_score=95.0 ピンで RSS(~85)に勝つ"],
            ["トレンド式", "score = (0.40×recency + 0.35×social + 0.25×authority) × 100"],
            ["recency", "指数減衰 exp(-ln2 × age_h / 48)。半減期48時間"],
            ["social", "0.7×log1p(upvotes)/log1p(10000) + 0.3×log1p(replies)/log1p(2000)。log で小規模も反映"],
            ["authority", "ソース別固定値。arxiv0.90 / r/ML0.75 / r/programming0.70 / bluesky0.70 / 既定0.50"],
            ["編集オーバーライド", "既存 trend_score(int/float)があれば再計算せず保持(2026-04-23、常緑シードのため)"],
            ["プラットフォーム割当", "RSS feed config の target、source_category(韓国/グルメ/コーヒー/AI等)で"
             "構成パターンへ"],
            ["クールダウン", "knowledge_topics は cooldown_days(既定30)で回転。rotation_weight=0 で真に無効化"],
        ],
        [22, 92],
    )

    # ---- 13 価格・ハッシュタグ --------------------------------------
    add_sheet(
        wb, "13_価格ハッシュタグ",
        "note 価格決定 & ハッシュタグ",
        "2026-05-12 メンバーシップ+有料審査通過で全記事有料化。価格は grade×evidence。",
        ["項目", "値 / 内容"],
        [
            ["価格 A+A", "¥1,980 (プレミアム)"],
            ["価格 A+B", "¥980"],
            ["価格 B+A", "¥500"],
            ["価格 B+B", "¥300"],
            ["価格 C以下", "¥200 (メンバーシップ誘導の床。note最低は¥100)"],
            ["価格関数", "NotePublisher.determine_price(overall_grade, evidence_level)"],
            ["借用画像オーバーライド", "本文に借用画像帰属マーカーがあれば price→¥0 強制(商用利用回避)"],
            ["ハッシュタグ", "HashtagGenerator(max_tags=5)。2026-05-28 に10→5へ(各タグのランキングsignal強化)"],
            ["フォールバックタグ", "生成失敗時 ['AI','テクノロジー','トレンド']"],
            ["cadence cap", "note 1本/日(NOTE_CADENCE_CAP, 真実源 data/publish_history.jsonl JST)。"
             "超過行は ✅承認 のまま翌日持ち越し"],
        ],
        [24, 90],
    )

    # ---- 14 主要パラメータ一覧 --------------------------------------
    add_sheet(
        wb, "14_主要パラメータ",
        "主要パラメータ一覧 (横断クイックリファレンス)",
        "値はコード/.env 探索時点。区分: ランタイム(.env) / コード default / 設定。",
        ["パラメータ", "値", "区分 / ファイル"],
        [
            ["LLM_MODEL_WRITER / SCORER", "gemma4:e4b", "ランタイム .env"],
            ["DEFAULT_MODEL", "gemma3:12b", "コード local_llm.py"],
            ["OLLAMA_API_URL", "http://localhost:11434", "ランタイム .env"],
            ["LLM timeout", "900s", "コード local_llm.py"],
            ["num_ctx", "16384", "コード local_llm.py"],
            ["温度(生成/評価/要約/コード)", "0.7 / 0.3 / 0.5 / 0.2", "運用ガイド"],
            ["RAG_ENABLED", "true", "ランタイム .env"],
            ["埋め込みモデル", "multilingual-e5-base (768d)", "build_rag_index.py"],
            ["ハルシ ガード閾値", "0.85", "main.py RAG"],
            ["学習ブロック取得閾値", "0.55 (top_k3)", "main.py RAG"],
            ["CHATGPT_CDP_PORT", "9222", "ランタイム .env"],
            ["vision-eval cutoff", "6 (再試行1)", "chatgpt_image_generator.py"],
            ["画像サイズ(既定)", "1792×1024 (16:9)", "chatgpt_image_generator.py"],
            ["最小画像バイト", "10,000", "chatgpt_batch_helper.py"],
            ["トレンド重み", "recency0.40 / social0.35 / authority0.25", "trend_detector.py"],
            ["recency 半減期", "48h", "trend_detector.py"],
            ["客観 evidence", "A≥80% / B60-79% / C<60%", "objective_scorer.py"],
            ["客観 visual_count", "A≥5 / B1-4 / C0", "objective_scorer.py"],
            ["客観 word_count 許容", "1,700-8,000", "objective_scorer.py"],
            ["主観 B+閾値", "2.0 (A=3/B=2/C=1)", "score_aggregator.py"],
            ["numeric_score", "0.5×客観 + 0.5×主観", "score_aggregator.py"],
            ["Zenn 記事閾値", "77.5", "main.py publish"],
            ["note 価格表", "1980/980/500/300/200", "note_publisher.determine_price"],
            ["note cadence cap", "1本/日", "main.py / .env NOTE_CADENCE_CAP"],
            ["ハッシュタグ数", "5", "HashtagGenerator"],
            ["週次トークン上限", "2,000,000", "settings.yaml"],
            ["Writer↔Critic round", "最大2(例外3)", "settings.yaml"],
            ["最低独立ソース", "3 (Tier1-2)", "settings.yaml"],
        ],
        [30, 38, 38],
    )

    # ---- 15 ファイル対応表 ------------------------------------------
    add_sheet(
        wb, "15_ファイル対応表",
        "AI関連 主要ファイル / モジュール対応表",
        "AI内部に関わる主要ファイルの責務。",
        ["ファイル / モジュール", "責務"],
        [
            ["main.py", "全パイプライン駆動。生成系 Phase1-3、publish_approved(Phase4)、RAG配線、publish deny"],
            ["generators/local_llm.py", "Ollama クライアント。生成/コード生成、Codex フォールバック"],
            ["generators/llm_config.py", "タスク別モデル選択 (LLM_MODEL_<TASK>)"],
            ["generators/objective_scorer.py", "客観スコアリング(足切り) 9指標+加点2"],
            ["generators/subjective_evaluator.py", "主観スコアリング(LLM) 5次元、ハルシ警告連携"],
            ["generators/score_aggregator.py", "総合判定・numeric_score・Sheets行生成"],
            ["generators/content_sanitizer.py", "スコアリング前のプレースホルダ/AI開示除去"],
            ["generators/rag_retriever.py", "RAG 検索(bi-encoder e5 + 任意 reranker)"],
            ["scripts/build_rag_index.py", "chromadb 索引構築(7コレクション)"],
            ["generators/chatgpt_image_generator.py", "ChatGPT画像生成(CDP/プロンプト/vision-eval/soft-delete)"],
            ["generators/visual_prompt_builder.py", "Gemma で記事→日本語ビジュアルプロンプト"],
            ["generators/image_style_presets.py", "スタイルプリセット(kbeauty_poster 等)"],
            ["generators/chatgpt_batch_helper.py", "画像バッチ + フォールバック段(Pollinations/Unsplash)"],
            ["generators/image_sourcer.py", "Unsplash/Pexels ストック画像(CC0系のみ)"],
            ["generators/hashtag_generator.py", "ハッシュタグ生成(max5)"],
            ["collectors/trend_detector.py", "トレンドスコア(recency/social/authority)"],
            ["collectors/*", "arXiv/Reddit/RSS/Bluesky/GoogleTrends/knowledge_topics 収集"],
            ["config/prompts.yaml", "全プロンプト+構成パターン+選択ルール"],
            ["config/settings.yaml", "禁止句40+/チェーン店BL/トークン上限/エージェント制約"],
            ["publishers/note_publisher.py", "note投稿(Selenium)・価格決定・edit_article"],
            ["publishers/zenn_publisher.py", "Zenn(git push) / zenn_scrap_publisher.py(scrap)"],
            ["AGENTS.md / .claude/skills/", "5エージェント設計・スキル定義"],
            ["docs/knowledge/hallucination_registry.md", "ハルシ事故レジストリ(正典)"],
            ["docs/knowledge/ops_incidents.md", "運用事象レジストリ(RAG ingest)"],
        ],
        [38, 78],
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"saved: {OUT}  ({len(wb.sheetnames)} sheets)")
    print("sheets:", wb.sheetnames)


if __name__ == "__main__":
    build()
